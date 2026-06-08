#!/usr/bin/env python3
"""
Gera a planilha de gastos da LP4 (Onde mora seu dinheiro).
Saída: ../public/guias/planilha-onde-mora-seu-dinheiro.xlsx
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, NamedStyle, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

INK = "FF233853"
INK_DEEP = "FF1a2c44"
PAPER = "FFFAF7F0"
PAPER_RAISED = "FFEEE9DD"
RULE = "FFD8D3C8"
MUTED = "FF6B7892"
CLAY = "FFB86B3A"

BRL_FMT = 'R$ #,##0;[Red]-R$ #,##0;"—"'
PCT_FMT = "0.0%;[Red]-0.0%"

OUT_DIR = Path(__file__).resolve().parent.parent / "public" / "guias"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT = OUT_DIR / "planilha-onde-mora-seu-dinheiro.xlsx"


def thin(color=RULE):
    return Side(border_style="thin", color=color)


def fill(color):
    return PatternFill("solid", fgColor=color)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Onde mora seu dinheiro"

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 4

    # ── HEADER ─────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    ws["B2"] = "§ MIDLEJ CAPITAL · MENTORIA"
    ws["B2"].font = Font(name="Calibri", size=9, bold=True, color=CLAY)
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[3].height = 8

    ws["B4"] = "Onde mora seu dinheiro"
    ws["B4"].font = Font(name="Calibri", size=26, bold=True, color=INK)
    ws["B4"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 40

    ws["B5"] = "Planilha estruturada de fluxo de caixa — médico PJ."
    ws["B5"].font = Font(name="Calibri", size=11, italic=True, color=MUTED)
    ws.row_dimensions[5].height = 18

    ws["B6"] = (
        "Preencha as caixas amarelas com os valores do seu mês típico. "
        "Os totais se movem sozinhos. O resultado em vermelho na linha 'Sobra' "
        "é o que está disponível para investir."
    )
    ws["B6"].font = Font(name="Calibri", size=10, color=INK)
    ws["B6"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells("B6:D6")
    ws.row_dimensions[6].height = 38

    ws.row_dimensions[7].height = 8

    # ── COLUMN HEADERS ─────────────────────────────────────
    ws["B8"] = "Categoria"
    ws["C8"] = "Detalhamento"
    ws["D8"] = "Valor mensal"
    for col in ("B", "C", "D"):
        cell = ws[f"{col}8"]
        cell.font = Font(name="Calibri", size=9, bold=True, color=MUTED)
        cell.alignment = Alignment(horizontal="left" if col != "D" else "right",
                                   vertical="center")
        cell.border = Border(bottom=Side(border_style="medium", color=INK))
    ws.row_dimensions[8].height = 22

    # ── SECTIONS ───────────────────────────────────────────
    ENTRADAS = [
        ("Salário PJ (pro-labore + lucros)", "líquido depois de impostos PJ"),
        ("Plantões e sobreaviso", "média dos últimos 6 meses"),
        ("Consultórios / particular", "valor recebido em conta no mês típico"),
        ("Outras rendas", "alugueres, dividendos, royalties"),
    ]
    FIXOS = [
        ("Moradia", "aluguel/condomínio/IPTU/contas"),
        ("Carro", "parcela, seguro, manutenção, combustível"),
        ("Saúde da família", "plano de saúde, terapias, academia"),
        ("Educação", "escola dos filhos, cursos, mensalidades"),
        ("Funcionários / clínica", "salários, aluguel da sala, materiais"),
    ]
    VARIAVEIS = [
        ("Alimentação", "mercado + restaurantes + delivery"),
        ("Lazer e viagens", "rateio mensal das viagens do ano"),
        ("Compras pessoais", "roupas, gadgets, presentes"),
        ("Outros e imprevistos", "o que não cabe em lugar nenhum"),
    ]

    DEFAULT_VALUES = {
        "Salário PJ (pro-labore + lucros)": 28000,
        "Plantões e sobreaviso": 12000,
        "Consultórios / particular": 8000,
        "Outras rendas": 0,
        "Moradia": 8000,
        "Carro": 3500,
        "Saúde da família": 2500,
        "Educação": 4500,
        "Funcionários / clínica": 5000,
        "Alimentação": 3500,
        "Lazer e viagens": 5000,
        "Compras pessoais": 2500,
        "Outros e imprevistos": 2000,
    }

    row = 9

    def write_section(title, items, sign):
        nonlocal row
        # Section title
        ws.cell(row=row, column=2, value=f"§ {title.upper()}").font = Font(
            name="Calibri", size=9, bold=True, color=CLAY
        )
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center")
        for col in range(2, 5):
            ws.cell(row=row, column=col).fill = fill(PAPER_RAISED)
        ws.row_dimensions[row].height = 22
        start = row + 1
        row += 1
        # Items
        for label, hint in items:
            ws.cell(row=row, column=2, value=label).font = Font(name="Calibri", size=11, color=INK)
            ws.cell(row=row, column=3, value=hint).font = Font(
                name="Calibri", size=9, italic=True, color=MUTED
            )
            ws.cell(row=row, column=4, value=DEFAULT_VALUES.get(label, 0)).number_format = BRL_FMT
            ws.cell(row=row, column=4).font = Font(name="Calibri", size=11, color=INK)
            ws.cell(row=row, column=4).fill = fill("FFFFF8DC")  # amarelo input
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
            for col in range(2, 5):
                ws.cell(row=row, column=col).border = Border(
                    bottom=Side(border_style="thin", color=RULE)
                )
            ws.row_dimensions[row].height = 22
            row += 1
        end = row - 1
        # Subtotal
        ws.cell(row=row, column=2, value=f"Subtotal — {title}").font = Font(
            name="Calibri", size=10, bold=True, color=INK
        )
        ws.cell(row=row, column=4, value=f"=SUM(D{start}:D{end})").number_format = BRL_FMT
        ws.cell(row=row, column=4).font = Font(name="Calibri", size=11, bold=True, color=INK)
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
        for col in range(2, 5):
            ws.cell(row=row, column=col).border = Border(top=Side(border_style="thin", color=INK))
        ws.row_dimensions[row].height = 26
        subtotal_row = row
        row += 1
        # spacer
        ws.row_dimensions[row].height = 14
        row += 1
        return subtotal_row, sign

    entradas_row, _ = write_section("Entradas", ENTRADAS, 1)
    fixos_row, _ = write_section("Custos fixos", FIXOS, -1)
    var_row, _ = write_section("Variáveis", VARIAVEIS, -1)

    # ── BOTTOM LINE ─────────────────────────────────────
    row += 1
    ws.cell(row=row, column=2, value="NO FIM DO MÊS").font = Font(
        name="Calibri", size=10, bold=True, color=PAPER
    )
    ws.cell(row=row, column=3, value="SOBRA REAL").font = Font(
        name="Calibri", size=9, color=MUTED
    )
    ws.cell(
        row=row,
        column=4,
        value=f"=D{entradas_row}-D{fixos_row}-D{var_row}",
    ).number_format = BRL_FMT
    ws.cell(row=row, column=4).font = Font(name="Calibri", size=14, bold=True, color=CLAY)
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
    for col in range(2, 5):
        ws.cell(row=row, column=col).fill = fill(INK)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 36
    sobra_row = row
    row += 1

    ws.cell(row=row, column=2, value="% da renda").font = Font(
        name="Calibri", size=9, color=MUTED
    )
    ws.cell(
        row=row, column=4, value=f"=D{sobra_row}/D{entradas_row}"
    ).number_format = PCT_FMT
    ws.cell(row=row, column=4).font = Font(name="Calibri", size=10, color=MUTED)
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
    ws.row_dimensions[row].height = 18

    row += 2

    # ── COMENTÁRIO DO AUDITOR ─────────────────────────────
    ws.cell(row=row, column=2, value="§ COMENTÁRIO DO AUDITOR").font = Font(
        name="Calibri", size=9, bold=True, color=CLAY
    )
    row += 1
    notas = [
        "Sobra acima de 20% — está parado em algum lugar. Você tem capacidade ociosa de aporte.",
        "Sobra entre 5% e 20% — apertado, mas saudável. O gargalo costuma estar em ‘variáveis’.",
        "Sobra abaixo de 5% — não está sobrando, está vazando. Hora de revisar custo fixo.",
        "Sobra negativa — você está consumindo patrimônio acumulado. Urgente revisar.",
    ]
    for n in notas:
        ws.cell(row=row, column=2, value=n).font = Font(name="Calibri", size=10, color=INK)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1
    footer = (
        "Midlej Capital · Mentoria — material informativo. Não constitui "
        "recomendação personalizada de investimentos."
    )
    ws.cell(row=row, column=2, value=footer).font = Font(
        name="Calibri", size=8.5, italic=True, color=MUTED
    )
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

    wb.save(OUT)
    print(f"✓ {OUT}")


if __name__ == "__main__":
    main()
